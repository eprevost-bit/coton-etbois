import base64
import io
import logging
from odoo import models, fields, _, api
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None


class PurchaseImportWizard(models.TransientModel):
    _name = 'purchase.import.wizard'
    _description = 'Asistente para Importar Precios de Compra'

    file_data = fields.Binary(string='Archivo Excel', required=True)
    file_name = fields.Char(string='Nombre del archivo')
    purchase_id = fields.Many2one('purchase.order', string='Pedido de Compra')

    def action_import_lines(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError("La librería openpyxl no está instalada.")

        # 1. Leer el archivo Excel
        try:
            file_content = base64.b64decode(self.file_data)
            data = io.BytesIO(file_content)
            workbook = openpyxl.load_workbook(data, data_only=True)
            sheet = workbook.active
        except Exception as e:
            raise UserError(f"Error al leer el archivo: {e}")

        # ---------------------------------------------------------
        # 2. MAPEO DINÁMICO DE COLUMNAS (¡LA SOLUCIÓN!)
        # ---------------------------------------------------------
        # Leemos la primera fila (Cabeceras) para saber dónde está cada cosa
        header_iterator = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        try:
            headers = next(header_iterator)  # Obtenemos la lista de nombres ['id', 'name', ...]
        except StopIteration:
            raise UserError("El archivo Excel parece estar vacío.")
        col_map = {str(h).strip(): i for i, h in enumerate(headers) if h}

        _logger.info(f"DEBUG - Columnas encontradas: {col_map}")

        # Verificamos que exista la columna OBLIGATORIA
        if 'order_line/id' not in col_map:
            raise UserError(
                f"No se encuentra la columna 'order_line/id' en el Excel.\n"
                f"Columnas detectadas: {list(col_map.keys())}"
            )

        # Obtenemos los índices (números de columna)
        idx_xml_id = col_map['order_line/id']
        # Usamos .get() por si acaso no existen, que devuelva None
        idx_price = col_map.get('order_line/price_unit')
        idx_qty = col_map.get('order_line/product_qty')

        # ---------------------------------------------------------
        # 3. RECORRER Y ACTUALIZAR
        # ---------------------------------------------------------
        updated_count = 0
        skipped_count = 0

        # Empezamos en la fila 2 (datos)
        for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

            # Obtener el ID usando el índice dinámico
            line_xml_id = row[idx_xml_id]

            if not line_xml_id:
                _logger.info(f"Fila {i}: Saltada (ID vacío)")
                skipped_count += 1
                continue

            # Limpiar el ID (quitar espacios)
            line_xml_id = str(line_xml_id).strip()

            # Buscar en Odoo
            line_record = self.env.ref(line_xml_id, raise_if_not_found=False)

            if not line_record:
                # Intento extra: A veces el ID viene sin el módulo "__export__"
                # Si el usuario subió un ID manual, esto ayuda a depurar
                _logger.warning(f"Fila {i}: ID '{line_xml_id}' NO ENCONTRADO en base de datos.")
                skipped_count += 1
                continue

            if line_record._name != 'purchase.order.line':
                _logger.warning(f"Fila {i}: El ID corresponde a '{line_record._name}', no a una línea de compra.")
                skipped_count += 1
                continue

            # Preparar valores
            vals = {}

            # -- PRECIO --
            if idx_price is not None:
                raw_price = row[idx_price]
                if raw_price is not None:
                    try:
                        vals['price_unit'] = float(raw_price)
                    except ValueError:
                        pass  # No era un número

            # -- CANTIDAD --
            if idx_qty is not None:
                raw_qty = row[idx_qty]
                if raw_qty is not None:
                    try:
                        vals['product_qty'] = float(raw_qty)
                    except ValueError:
                        pass

            # Escribir
            if vals:
                line_record.write(vals)
                updated_count += 1
                _logger.info(f"Fila {i}: Actualizada OK ({line_xml_id}) -> {vals}")
            else:
                skipped_count += 1
                _logger.info(f"Fila {i}: Sin cambios detectados o datos vacíos.")

        # ---------------------------------------------------------
        # 4. RESULTADO
        # ---------------------------------------------------------
        return {
            'type': 'ir.actions.client',
            'tag': 'display_notification',
            'params': {
                'title': _('Proceso Terminado'),
                'message': f'✅ Actualizadas: {updated_count}\n❌ Omitidas/No encontradas: {skipped_count}',
                'type': 'success' if updated_count > 0 else 'warning',
                'sticky': False,
                'next': {'type': 'ir.actions.act_window_close'},
            }
        }